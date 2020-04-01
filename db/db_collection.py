import os
from openpyxl import load_workbook
from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dreamCard                    # 'dreamCard'라는 이름의 db를 만듭니다.

cwd = os.getcwd()  # Get the current working directory (cwd)
files_collection = os.listdir(cwd)  # Get all the files in that directory
xlsx_files = [file for file in files_collection if file.endswith('.xlsx')]

# for files in files_collection:
#     print(files)

num =0
for county_file_name in xlsx_files:
    file_name = cwd + '/' + county_file_name  # 맥, 윈도우 계속 수정해야한다.
    work_book = load_workbook(file_name, data_only=True)  # 파일의 제목을 읽어오는 작업
    work_sheet = work_book['SHEET1']  # 엑셀 파일을 열었을때 안에 있는 정보들을 읽어와라
    county_name = county_file_name.split('.')[1]

    for col in work_sheet.iter_rows(min_row=4, max_row=100000):

        if col[1].value is not None:
            restaurant_name = col[1].value
        else:
            continue
        if col[2].value is not None:
            restaurant_type = col[2].value
        else:
            continue
        if col[3].value is not None:
            restaurant_phoneNumber = col[3].value
        else:
            continue
        if col[4].value is not None:
            restaurant_address = col[4].value
        else:
            continue


        # restaurant_name = col[1].value if col[1].value is not None else '0'
        # restaurant_type = col[2].value if col[2].value is not None else '0'
        # restaurant_phoneNumber = col[3].value if col[3].value is not None else '0'
        # restaurant_address = col[4].value if col[4].value is not None else '0'

        restaurants_infos = {
            'restaurant_name': restaurant_name,
            'restaurant_type': restaurant_type,
            'restaurant_phoneNumber': restaurant_phoneNumber,
            'restaurant_address': restaurant_address
        }
        num += 1
        db.county_files.insert_one(restaurants_infos) #정보들 insert 하기
        #print(num, restaurants_infos)
        #print(col[1].value, col[2].value, col[3].value, col[4].value)
        print(num, col[1].value,col[4].value)
